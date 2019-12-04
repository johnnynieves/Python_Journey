import openpyxl as xl
from openpyxl.chart import BarChart, Reference


wb = xl.load_workbook('transactions.xlsx')
# Workbook object
sheet = wb['Sheet1']  # I think this is sheet name
'''
# access exam1
cell = sheet['a1']

# access exam 2 method better
cell = sheet.cell(1, 1)
# print value of cell
# if you try to print just cell you will get a reference (<Cell 'Sheet1'.A1>)
print(cell.value)

# find out how many rows in sheet
print(sheet.max_row)
'''
# Dont use heading just use 2 to skip looking for it
rows = sheet.max_row + 1  # +1 is to include the last row in the range
for row in range(2, rows):
    cell = sheet.cell(row, 3)  # finds the money cell
    corrected_price = cell.value * 0.9  # corrective action
    corrected_price_cell = sheet.cell(row, 4)  # reference row and column
    # insert corrective action value into the cell
    corrected_price_cell.value = corrected_price

# Make BarChart
# below is how big the barchart is going to be
values = Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4
                   )

# Create the barchart based on the above values
chart = BarChart()
chart.add_data(values)

# Add the chart to sheet
sheet.add_chart(chart, 'e2')


wb.save('transactions2.xlsx')
