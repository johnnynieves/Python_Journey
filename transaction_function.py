import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']  # I think this is sheet name
    rows = sheet.max_row + 1  # +1 is to include the last row in the range
    for row in range(2, rows):
        cell = sheet.cell(row, 3)  # finds the money cell
        corrected_price = cell.value * 0.9  # corrective action
        corrected_price_cell = sheet.cell(row, 4)  # reference row and column
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4
                       )

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save(filename)


if __name__ == "__main__":
    process_workbook('transactions.xlsx')
